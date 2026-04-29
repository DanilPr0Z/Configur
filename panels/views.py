from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from collections import Counter
import openpyxl
from io import BytesIO
from datetime import datetime, date

from .models import (
    JointType, FinishGroup, Finish, ProfileColor,
    AluminumProfile, Order, DoorPanel, Panel,
)
from .serializers import (
    JointTypeSerializer, FinishGroupSerializer, ProfileColorSerializer,
    AluminumProfileSerializer, OrderListSerializer, OrderDetailSerializer,
    PanelSerializer, DoorPanelSerializer,
)


class JointTypeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = JointType.objects.all()
    serializer_class = JointTypeSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx['request'] = self.request
        return ctx

    @action(detail=True, methods=['post'], url_path='upload-image')
    def upload_image(self, request, pk=None):
        joint = get_object_or_404(JointType, pk=pk)
        file = request.FILES.get('image')
        if not file:
            return Response({'error': 'Файл не передан'}, status=status.HTTP_400_BAD_REQUEST)
        if joint.image:
            joint.image.delete(save=False)
        joint.image.save(file.name, file, save=True)
        serializer = self.get_serializer(joint)
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='delete-image')
    def delete_image(self, request, pk=None):
        joint = get_object_or_404(JointType, pk=pk)
        if joint.image:
            joint.image.delete(save=True)
        return Response({'ok': True})


class FinishGroupViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FinishGroup.objects.prefetch_related('finishes').all()
    serializer_class = FinishGroupSerializer


class ProfileColorViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ProfileColor.objects.all()
    serializer_class = ProfileColorSerializer


class AluminumProfileViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AluminumProfile.objects.all()
    serializer_class = AluminumProfileSerializer


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()

    def get_serializer_class(self):
        if self.action in ('retrieve', 'create', 'update', 'partial_update'):
            return OrderDetailSerializer
        return OrderListSerializer

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        """Итоговая спецификация заказа с алюминиевыми профилями"""
        order = get_object_or_404(Order, pk=pk)
        panels = list(order.panels.select_related(
            'joint_left', 'joint_right', 'joint_top', 'joint_bottom',
            'finish', 'finish_group', 'aluminum_color',
        ).all())

        # Подсчёт использований типов узлов
        joint_counter = Counter()
        for panel in panels:
            for joint in [panel.joint_left, panel.joint_right]:
                if joint and joint.code:
                    joint_counter[joint.code] += panel.quantity
            for joint in [panel.joint_top, panel.joint_bottom]:
                if joint and joint.code:
                    joint_counter[joint.code] += panel.quantity

        # Автоподбор профилей
        profiles = AluminumProfile.objects.all()
        profile_specs = []
        for profile in profiles:
            if profile.joint_type_code:
                count_joints = joint_counter.get(profile.joint_type_code, 0)
                qty = round(count_joints * profile.count_per_joint)
            elif profile.article == 'МДФ 10':
                total_panels = sum(p.quantity for p in panels)
                qty = total_panels * 4
            else:
                qty = 0

            if qty > 0:
                profile_specs.append({
                    'id': profile.id,
                    'article': profile.article,
                    'name': profile.name,
                    'length_mm': profile.length_mm,
                    'quantity': qty,
                    'price_per_piece': profile.price_per_piece,
                    'total_cost': qty * profile.price_per_piece,
                    'note': profile.note,
                })

        profiles_total = sum(p['total_cost'] for p in profile_specs)
        panels_total = sum(p.total_cost for p in panels)
        door_panels_total = sum(
            d.total_cost for d in order.door_panels.select_related('finish').all()
        )

        return Response({
            'order_id': order.id,
            'customer_name': order.customer_name,
            'order_number': order.order_number,
            'panels_count': len(panels),
            'panels_total': round(panels_total, 2),
            'door_panels_total': round(door_panels_total, 2),
            'profiles': profile_specs,
            'profiles_total': round(profiles_total, 2),
            'grand_total': round(panels_total + door_panels_total + profiles_total, 2),
        })

    @action(detail=True, methods=['post'])
    def import_excel(self, request, pk=None):
        """
        Импорт панелей из Excel-шаблона NUOVO 60.
        POST multipart/form-data с полем file (xlsx).
        Читает лист «Ввод данных к заказу», строки 145-174.
        """
        order = get_object_or_404(Order, pk=pk)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Файл не передан'}, status=400)

        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        except Exception:
            return Response({'error': 'Не удалось открыть файл. Убедитесь, что это xlsx.'}, status=400)

        sheet_name = 'Ввод данных к заказу'
        if sheet_name not in wb.sheetnames:
            return Response({'error': f'Лист "{sheet_name}" не найден в файле'}, status=400)

        ws = wb[sheet_name]

        def cell_str(row, col):
            v = ws.cell(row=row, column=col).value
            if v is None:
                return ''
            s = str(v).strip()
            return '' if s in ('#N/A', '#VALUE!', '#REF!') else s

        def cell_num(row, col, default=0):
            v = ws.cell(row=row, column=col).value
            try:
                return float(v) if v is not None else default
            except (ValueError, TypeError):
                return default

        # --- Обновляем данные заказа из шапки (col E = 5) ---
        order_fields = {}
        for attr, row in [('customer_name', 62), ('agent_name', 64),
                          ('counterparty', 65), ('order_number', 66),
                          ('invoice_number', 68), ('city', 72)]:
            val = cell_str(row, 5)
            if val:
                order_fields[attr] = val

        date_val = ws.cell(row=70, column=5).value
        if date_val:
            if isinstance(date_val, (datetime, date)):
                order_fields['order_date'] = date_val.date() if isinstance(date_val, datetime) else date_val
            elif isinstance(date_val, str):
                for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
                    try:
                        order_fields['order_date'] = datetime.strptime(date_val, fmt).date()
                        break
                    except ValueError:
                        pass

        if order_fields:
            for f, v in order_fields.items():
                setattr(order, f, v)
            order.save()

        # --- Справочники ---
        joint_map = {j.code.upper().strip(): j for j in JointType.objects.all()}

        finish_group_map = {}
        finish_map = {}
        for fg in FinishGroup.objects.prefetch_related('finishes').all():
            finish_group_map[fg.name.upper().strip()] = fg
            for f in fg.finishes.all():
                finish_map[f.name.upper().strip()] = f

        color_map = {c.name.upper().strip(): c for c in ProfileColor.objects.all()}

        # --- Удаляем существующие панели и импортируем новые ---
        order.panels.all().delete()
        created = 0

        for row_num in range(145, 175):
            height = cell_num(row_num, 3)
            width = cell_num(row_num, 5)
            if not height and not width:
                continue

            jl_code = cell_str(row_num, 4).upper()
            jr_code = cell_str(row_num, 6).upper()
            qty_raw = cell_num(row_num, 7, 1)
            quantity = max(1, int(qty_raw) if qty_raw else 1)
            jt_code = cell_str(row_num, 9).upper()
            jb_code = cell_str(row_num, 10).upper()
            fg_name = cell_str(row_num, 12).upper()
            f_name = cell_str(row_num, 13).upper()
            decor = cell_str(row_num, 14)
            al_vert = int(cell_num(row_num, 15))
            al_horiz = int(cell_num(row_num, 16))
            al_color_name = cell_str(row_num, 17).upper()
            markup = cell_num(row_num, 19)

            Panel.objects.create(
                order=order,
                position=created + 1,
                height_mm=height,
                width_mm=width,
                quantity=quantity,
                joint_left=joint_map.get(jl_code),
                joint_right=joint_map.get(jr_code),
                joint_top=joint_map.get(jt_code),
                joint_bottom=joint_map.get(jb_code),
                finish_group=finish_group_map.get(fg_name),
                finish=finish_map.get(f_name),
                decor_name=decor,
                aluminum_vertical_count=al_vert,
                aluminum_horizontal_count=al_horiz,
                aluminum_color=color_map.get(al_color_name),
                markup_percent=markup,
            )
            created += 1

        return Response({'panels_imported': created, 'order_updated': bool(order_fields)})

    @action(detail=False, methods=['post'])
    def calculate_wall(self, request):
        """
        Калькулятор раскладки: считает ширины панелей из параметров отрезка стены.
        POST body:
          wall_length, panel_count, joint_left_code, joint_right_code, connection_type_code
        """
        wall_length = float(request.data.get('wall_length', 0))
        panel_count = int(request.data.get('panel_count', 1))
        joint_left_code = request.data.get('joint_left_code', '')
        joint_right_code = request.data.get('joint_right_code', '')
        connection_code = request.data.get('connection_type_code', 'B')

        try:
            jl = JointType.objects.get(code=joint_left_code)
        except JointType.DoesNotExist:
            return Response({'error': f'Узел {joint_left_code!r} не найден'}, status=400)
        try:
            jr = JointType.objects.get(code=joint_right_code)
        except JointType.DoesNotExist:
            return Response({'error': f'Узел {joint_right_code!r} не найден'}, status=400)

        if panel_count <= 0:
            return Response({'error': 'panel_count должен быть > 0'}, status=400)

        # Формула из Excel F131:
        # Длина_по_панелям = стена + offset_лев + offset_прав - (если_C: (n-1)*4)
        correction = (panel_count - 1) * 4 if connection_code == 'C' else 0
        total_panel_length = wall_length + jl.offset_mm + jr.offset_mm - correction
        panel_width = round(total_panel_length / panel_count, 1)

        return Response({
            'wall_length': wall_length,
            'total_panel_length': round(total_panel_length, 1),
            'panel_count': panel_count,
            'panel_width': panel_width,
            'joint_left': joint_left_code,
            'joint_right': joint_right_code,
        })


class PanelViewSet(viewsets.ModelViewSet):
    serializer_class = PanelSerializer

    def get_queryset(self):
        qs = Panel.objects.select_related(
            'joint_left', 'joint_right', 'joint_top', 'joint_bottom',
            'finish', 'finish_group', 'aluminum_color',
        )
        order_id = self.request.query_params.get('order')
        if order_id:
            qs = qs.filter(order_id=order_id)
        return qs


class DoorPanelViewSet(viewsets.ModelViewSet):
    serializer_class = DoorPanelSerializer

    def get_queryset(self):
        qs = DoorPanel.objects.select_related(
            'joint_top_left', 'joint_top_right', 'joint_bottom',
            'finish', 'finish_group',
        )
        order_id = self.request.query_params.get('order')
        if order_id:
            qs = qs.filter(order_id=order_id)
        return qs
