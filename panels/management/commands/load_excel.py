"""
Загрузка данных из Excel-файла NUOVO 60 в базу данных.
Запуск: python manage.py load_excel
"""

import os
from django.core.management.base import BaseCommand
import openpyxl

from panels.models import (
    JointType, FinishGroup, Finish, ProfileColor, AluminumProfile,
)

EXCEL_PATH = 'NUOVO 60. Расчет стеновых панелей (учет узлов) 26.11.2025.xlsx'

# Коды узлов
JOINT_CODES = {'A', 'B', 'C', 'D', 'DG', 'DH', 'E', 'FL', 'FR', 'G', 'H', 'I', 'J', 'K', 'O', 'P', 'R', 'S', 'T'}

# Маппинг col F -> название группы отделок
GROUP_COL_MAP = {
    'ШПОН': 'ШПОН',
    'FRASSINO': 'FRASSINO',
    'КОМПОЗИТ': 'КОМПОЗИТ',
    'LACATO': 'LACATO',
    'ШПОН 5 ММ': 'ШПОН 5 ММ',
    'ШПОН 2,5 ММ': 'ШПОН 2,5 ММ',
    'ШПОН 1,5 ММ': 'ШПОН 1,5 ММ',
    'FONDO': 'FONDO',
    'MIRROR': 'MIRROR',
    'Colour Glass': 'COLOUR GLASS',
    'Colour Glass mat': 'COLOUR GLASS MAT',
    'Colour Glass 8W': 'COLOUR GLASS 8W',
    'Colour Glass mat 8W': 'COLOUR GLASS MAT 8W',
    'LACATO_2.5_MM': 'LACATO 2,5 ММ',
}

# Названия камней из группы STONE
STONE_STARTS = ('Marmo', 'Rame FB', 'Ciliegia', 'Penelope', 'Linen', 'Skin Late', 'Pietra', 'Bronzo')

# Строки, которые не являются названием отделки
SKIP_NAMES = {
    'ШПОН', 'FRASSINO', 'LACATO', 'FONDO', 'GLOSS', 'STONE',
    'МАТЕРИАЛЫ', 'ОТДЕЛКИ', 'отделка', 'РАСХОДЫ ПАНЕЛИ', 'СЛОЕВ',
}

NON_COLOR_NAMES = {
    'Цвет профиля ', 'Группа отделок', 'ШПОН', 'STONE', 'LACATO',
    'FONDO', 'GLOSS', 'FRASSINO',
}


def norm_article(val):
    """Нормализует артикул профиля."""
    if val is None:
        return ''
    if isinstance(val, float):
        if abs(val - 104.256) < 0.001:
            return '104.256'
        if abs(val - 104.259) < 0.001:
            return '104.259'
        if abs(val - 104.27) < 0.001:
            return '104.270'
        return str(val)
    return str(val).strip()


def detect_group_by_name(name: str):
    """Определяет группу отделки по имени для случаев без маркера в col F."""
    if 'FRASSINO OLD' in name:
        return 'FRASSINO OLD'
    if '(PELLE)' in name or name.startswith('Pele '):
        return 'КОЖА'
    if name.startswith(STONE_STARTS):
        return 'STONE'
    if '(WOOD)' in name:
        return 'WOOD'
    return None


class Command(BaseCommand):
    help = 'Загружает данные из Excel-файла NUOVO 60 в базу данных'

    def handle(self, *args, **options):
        if not os.path.exists(EXCEL_PATH):
            self.stderr.write(f'Файл не найден: {EXCEL_PATH}')
            return

        self.stdout.write('Открываю Excel...')
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        self._load_joint_types(wb)
        self._load_profile_colors(wb)
        self._load_finishes(wb)
        self._load_aluminum_profiles(wb)

        self.stdout.write(self.style.SUCCESS('\n✓ Загрузка завершена!'))

    # ──────────────────────────────────────────────────────────────────────────
    # 1. Типы узлов — лист DATA
    # ──────────────────────────────────────────────────────────────────────────
    def _load_joint_types(self, wb):
        self.stdout.write('\n--- Типы узлов ---')
        ws = wb['DATA']

        # Собираем данные по узлам: код -> (offset, count, article, price)
        # Данные: idx3=код, idx4=offset, idx5=кол-во, idx6=артикул, idx7=цена
        joint_map = {}
        for row in ws.iter_rows(values_only=True):
            code = row[3] if len(row) > 3 else None
            if not isinstance(code, str):
                continue
            code = code.strip()
            if code not in JOINT_CODES:
                continue

            offset = float(row[4]) if isinstance(row[4], (int, float)) else 0.0
            count = float(row[5]) if isinstance(row[5], (int, float)) else 0.0
            article = norm_article(row[6]) if len(row) > 6 else ''
            price = float(row[7]) if len(row) > 7 and isinstance(row[7], (int, float)) else 0.0

            # Если код уже есть — обновляем только если новая запись богаче
            if code not in joint_map:
                joint_map[code] = (offset, count, article, price)
            else:
                old = joint_map[code]
                # Берём запись с ценой, либо с большим количеством данных
                if price > old[3] or (price == old[3] and article and not old[2]):
                    joint_map[code] = (offset, count, article, price)

        # Описания узлов
        labels = {
            'A': 'Торцевой (финиш)',
            'B': 'Ламель (соединение)',
            'C': 'Соединительный',
            'D': 'Угол наружный 90°',
            'DG': 'Угол внутренний G',
            'DH': 'Угол внутренний H',
            'E': 'Торцевой (+26 мм)',
            'FL': 'Финишный левый',
            'FR': 'Финишный правый',
            'G': 'Тип G (дверь, наружу)',
            'H': 'Тип H (дверь, внутрь)',
            'I': 'Тип I',
            'J': 'Тип J',
            'K': 'Тип K',
            'O': 'Без профиля',
            'P': 'Декор',
            'R': 'Подрез',
            'S': 'Стык',
            'T': 'Тип T',
        }

        created = updated = 0
        for code, (offset, count, article, price) in joint_map.items():
            obj, is_new = JointType.objects.update_or_create(
                code=code,
                defaults=dict(
                    name=labels.get(code, ''),
                    offset_mm=offset,
                    profile_count=count,
                    profile_article=article,
                    price_per_meter=price,
                ),
            )
            if is_new:
                created += 1
            else:
                updated += 1

        self.stdout.write(f'  Создано: {created}, обновлено: {updated}')
        for code, (offset, count, article, price) in sorted(joint_map.items()):
            self.stdout.write(f'  {code:4s}  offset={offset:>7.1f}  цена={price:>5.0f} руб/пм  профиль={article}')

    # ──────────────────────────────────────────────────────────────────────────
    # 2. Цвета профилей — лист DATA
    # ──────────────────────────────────────────────────────────────────────────
    def _load_profile_colors(self, wb):
        self.stdout.write('\n--- Цвета профилей ---')
        ws = wb['DATA']

        colors = []
        reached_groups = False
        for row in ws.iter_rows(values_only=True):
            val = row[1] if len(row) > 1 else None
            if not isinstance(val, str):
                continue
            val_stripped = val.strip()
            if val_stripped == 'Группа отделок':
                reached_groups = True
                break
            if val_stripped in NON_COLOR_NAMES or not val_stripped:
                continue
            colors.append(val_stripped)

        created = updated = 0
        for i, name in enumerate(colors):
            _, is_new = ProfileColor.objects.update_or_create(
                name=name,
                defaults={'sort_order': i + 1},
            )
            if is_new:
                created += 1
            else:
                updated += 1

        self.stdout.write(f'  Создано: {created}, обновлено: {updated}  ({len(colors)} цветов)')

    # ──────────────────────────────────────────────────────────────────────────
    # 3. Группы отделок и отделки — лист Отделки
    # ──────────────────────────────────────────────────────────────────────────
    def _load_finishes(self, wb):
        self.stdout.write('\n--- Отделки ---')
        ws = wb['Отделки']

        current_group_name = None
        group_sort = 0
        finish_count = 0
        group_counts = {}

        for row in ws.iter_rows(values_only=True):
            name_raw = row[1] if len(row) > 1 else None
            price_sqm = row[3] if len(row) > 3 else None   # ЦЕНА КВ.М отделка
            group_marker = row[5] if len(row) > 5 else None  # col F — маркер группы

            # Обновляем текущую группу по маркеру col F
            if group_marker is not None:
                mapped = GROUP_COL_MAP.get(str(group_marker).strip())
                if mapped:
                    current_group_name = mapped

            # Пропускаем строки без имени
            if name_raw is None:
                continue

            name = str(name_raw).strip()
            if not name or name in SKIP_NAMES:
                continue

            # Пытаемся определить группу по имени (для случаев без маркера)
            detected = detect_group_by_name(name)
            if detected:
                current_group_name = detected

            # Пропускаем строки без цены или с нечисловой ценой
            if not isinstance(price_sqm, (int, float)):
                continue
            if current_group_name is None:
                continue

            # Создаём группу если нужно
            if current_group_name not in group_counts:
                group_sort += 1
                FinishGroup.objects.get_or_create(
                    name=current_group_name,
                    defaults={'sort_order': group_sort},
                )
                group_counts[current_group_name] = 0

            group = FinishGroup.objects.get(name=current_group_name)
            _, is_new = Finish.objects.update_or_create(
                group=group,
                name=name,
                defaults={'price_sqm': round(float(price_sqm), 2)},
            )
            group_counts[current_group_name] += 1
            finish_count += 1

        self.stdout.write(f'  Всего отделок загружено: {finish_count}')
        for g, cnt in sorted(group_counts.items()):
            self.stdout.write(f'  {g}: {cnt} шт.')

    # ──────────────────────────────────────────────────────────────────────────
    # 4. Алюминиевые профили — по данным из Excel + логика узлов
    # ──────────────────────────────────────────────────────────────────────────
    def _load_aluminum_profiles(self, wb):
        self.stdout.write('\n--- Алюминиевые профили ---')

        # Читаем цены из листа Отделки: idx11=артикул, idx12=цена поставщика, idx13=наша цена
        ws = wb['Отделки']
        profile_prices = {}  # article -> our_price
        for row in ws.iter_rows(values_only=True):
            article_raw = row[11] if len(row) > 11 else None
            our_price = row[13] if len(row) > 13 else None
            if article_raw is None:
                continue
            article = norm_article(article_raw)
            if article and isinstance(our_price, (int, float)):
                profile_prices[article] = float(our_price)

        self.stdout.write(f'  Цены профилей из Excel: {profile_prices}')

        # Список профилей: (article, name, length_mm, joint_type_code, count_per_joint, note)
        profiles = [
            ('104.256', 'Торцевой профиль (финишный) Арт. 104.256', 3000, 'A', 1.0, ''),
            ('104.256', 'Торцевой профиль (финишный) Арт. 104.256', 3000, 'I', 1.0, ''),
            ('104.259', 'Соединительный профиль Арт. 104.259', 3000, 'C', 0.5, ''),
            ('104.270', 'Угловой профиль Арт. 104.270', 3000, 'D', 0.5, ''),
            ('104.270', 'Угловой профиль Арт. 104.270', 3000, 'DG', 0.5, ''),
            ('104.270', 'Угловой профиль Арт. 104.270', 3000, 'DH', 0.5, ''),
            ('ламель', 'Ламель соединительная (тип B)', 3000, 'B', 0.5, ''),
            ('МДФ 10', 'Навес стеновой панели', 900, '', 0.0, '4 шт на каждую панель'),
        ]

        # Удаляем старые записи и создаём заново
        AluminumProfile.objects.all().delete()

        for article, name, length, joint_code, count, note in profiles:
            price = profile_prices.get(article, 0.0)
            # ламель -> артикул 'ламель ' в Excel (с пробелом)
            if article == 'ламель':
                price = profile_prices.get('ламель ', profile_prices.get('ламель', 0.0))

            AluminumProfile.objects.create(
                article=article,
                name=name,
                length_mm=length,
                price_per_piece=price,
                joint_type_code=joint_code,
                count_per_joint=count,
                note=note,
            )
            self.stdout.write(f'  {article:12s} / {joint_code:4s}  цена={price:.0f} руб/шт  ({name[:40]})')
